"""
Attendance Report Generator Module
Module tạo báo cáo điểm danh toàn diện
"""

import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64

class AttendanceReportGenerator:
    def __init__(self, db_path: str):
        """
        Initialize report generator
        
        Args:
            db_path: Path to SQLite database file
        """
        self.db_path = db_path
        
    def get_connection(self):
        """Get database connection"""
        return sqlite3.connect(self.db_path)
    
    # =============================================================================
    # DATA RETRIEVAL METHODS
    # =============================================================================
    
    def get_attendance_by_date_range(self, start_date: str, end_date: str, class_id: Optional[int] = None) -> pd.DataFrame:
        """
        Get attendance records for date range
        
        Args:
            start_date: Start date (YYYY-MM-DD format)
            end_date: End date (YYYY-MM-DD format) 
            class_id: Optional class filter
            
        Returns:
            DataFrame with attendance records
        """
        query = """
        SELECT 
            date(ar.check_in_time) as date,
            time(ar.check_in_time) as time,
            s.full_name as student_name,
            s.student_id as student_code,
            c.class_name,
            ar.confidence_score,
            ass.session_name
        FROM attendance_records ar
        JOIN students s ON ar.student_id = s.id
        JOIN classes c ON s.class_id = c.id
        LEFT JOIN attendance_sessions ass ON ar.session_id = ass.id
        WHERE date(ar.check_in_time) BETWEEN ? AND ?
        """
        
        params = [start_date, end_date]
        
        if class_id:
            query += " AND s.class_id = ?"
            params.append(class_id)
            
        query += " ORDER BY ar.check_in_time DESC"
        
        with self.get_connection() as conn:
            return pd.read_sql_query(query, conn, params=params)
    
    def get_daily_attendance_summary(self, start_date: str, end_date: str) -> pd.DataFrame:
        """Get daily attendance summary statistics"""
        query = """
        SELECT 
            date(ar.check_in_time) as attendance_date,
            c.class_name,
            COUNT(DISTINCT ar.student_id) as present_count,
            COUNT(DISTINCT s.id) as total_students,
            ROUND(COUNT(DISTINCT ar.student_id) * 100.0 / COUNT(DISTINCT s.id), 2) as attendance_rate
        FROM attendance_records ar
        JOIN students s ON ar.student_id = s.id
        JOIN classes c ON s.class_id = c.id
        WHERE date(ar.check_in_time) BETWEEN ? AND ?
        GROUP BY date(ar.check_in_time), c.id, c.class_name
        ORDER BY attendance_date DESC, c.class_name
        """
        
        with self.get_connection() as conn:
            return pd.read_sql_query(query, conn, params=[start_date, end_date])
    
    def get_student_attendance_summary(self, start_date: str, end_date: str, student_id: Optional[int] = None) -> pd.DataFrame:
        """Get attendance summary by student"""
        query = """
        SELECT 
            s.full_name as student_name,
            s.student_id as student_code,
            c.class_name,
            COUNT(DISTINCT date(ar.check_in_time)) as days_present,
            COUNT(DISTINCT ass.id) as total_sessions,
            ROUND(COUNT(DISTINCT date(ar.check_in_time)) * 100.0 / 
                  (julianday(?) - julianday(?) + 1), 2) as attendance_percentage
        FROM students s
        JOIN classes c ON s.class_id = c.id
        LEFT JOIN attendance_records ar ON s.id = ar.student_id 
            AND date(ar.check_in_time) BETWEEN ? AND ?
        LEFT JOIN attendance_sessions ass ON date(ass.session_date) BETWEEN ? AND ?
        """
        
        params = [end_date, start_date, start_date, end_date, start_date, end_date]
        
        if student_id:
            query += " WHERE s.id = ?"
            params.append(student_id)
            
        query += " GROUP BY s.id, s.full_name, s.student_id, c.class_name"
        query += " ORDER BY attendance_percentage DESC"
        
        with self.get_connection() as conn:
            return pd.read_sql_query(query, conn, params=params)
    
    def get_class_statistics(self, class_id: Optional[int] = None) -> pd.DataFrame:
        """Get overall class statistics"""
        query = """
        SELECT 
            c.class_name,
            COUNT(s.id) as total_students,
            COUNT(DISTINCT ar.student_id) as students_with_attendance,
            COUNT(ar.id) as total_attendance_records,
            AVG(ar.confidence_score) as avg_confidence
        FROM classes c
        LEFT JOIN students s ON c.id = s.class_id
        LEFT JOIN attendance_records ar ON s.id = ar.student_id
        """
        
        params = []
        if class_id:
            query += " WHERE c.id = ?"
            params.append(class_id)
            
        query += " GROUP BY c.id, c.class_name ORDER BY c.class_name"
        
        with self.get_connection() as conn:
            return pd.read_sql_query(query, conn, params=params)
    
    def get_attendance_trends(self, days: int = 30) -> pd.DataFrame:
        """Get attendance trends over specified days"""
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        
        query = """
        SELECT 
            date(ar.check_in_time) as date,
            COUNT(DISTINCT ar.student_id) as unique_students,
            COUNT(ar.id) as total_records,
            AVG(ar.confidence_score) as avg_confidence
        FROM attendance_records ar
        WHERE date(ar.check_in_time) BETWEEN ? AND ?
        GROUP BY date(ar.check_in_time)
        ORDER BY date
        """
        
        with self.get_connection() as conn:
            return pd.read_sql_query(query, conn, params=[start_date, end_date])
    
    def get_hourly_attendance_pattern(self) -> pd.DataFrame:
        """Get attendance patterns by hour of day"""
        query = """
        SELECT 
            CAST(strftime('%H', ar.check_in_time) AS INTEGER) as hour,
            COUNT(ar.id) as attendance_count,
            COUNT(DISTINCT ar.student_id) as unique_students
        FROM attendance_records ar
        GROUP BY CAST(strftime('%H', ar.check_in_time) AS INTEGER)
        ORDER BY hour
        """
        
        with self.get_connection() as conn:
            return pd.read_sql_query(query, conn)
    
    # =============================================================================
    # EXCEL EXPORT METHODS
    # =============================================================================
    
    def create_styled_worksheet(self, workbook, sheet_name: str, data: pd.DataFrame, title: str):
        """Create a styled worksheet with data"""
        ws = workbook.create_sheet(title=sheet_name)
        
        # Add title
        ws.merge_cells('A1:' + chr(65 + len(data.columns) - 1) + '1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        
        # Add headers
        header_row = 3
        for col_idx, column in enumerate(data.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = column
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row_idx, row in enumerate(data.itertuples(index=False), start=header_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    def export_comprehensive_report(self, start_date: str, end_date: str, output_path: str) -> str:
        """
        Export comprehensive attendance report to Excel
        
        Args:
            start_date: Start date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD) 
            output_path: Output file path
            
        Returns:
            Path to created file
        """
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Get data
        attendance_data = self.get_attendance_by_date_range(start_date, end_date)
        daily_summary = self.get_daily_attendance_summary(start_date, end_date)
        student_summary = self.get_student_attendance_summary(start_date, end_date)
        class_stats = self.get_class_statistics()
        trends = self.get_attendance_trends()
        hourly_pattern = self.get_hourly_attendance_pattern()
        
        # Create sheets
        if not attendance_data.empty:
            self.create_styled_worksheet(
                wb, 
                "Chi tiết điểm danh", 
                attendance_data, 
                f"Chi tiết điểm danh ({start_date} đến {end_date})"
            )
        
        if not daily_summary.empty:
            self.create_styled_worksheet(
                wb, 
                "Tổng hợp theo ngày", 
                daily_summary, 
                f"Tổng hợp điểm danh theo ngày ({start_date} đến {end_date})"
            )
        
        if not student_summary.empty:
            self.create_styled_worksheet(
                wb, 
                "Tổng hợp học sinh", 
                student_summary, 
                f"Tổng hợp điểm danh theo học sinh ({start_date} đến {end_date})"
            )
        
        if not class_stats.empty:
            self.create_styled_worksheet(
                wb, 
                "Thống kê lớp học", 
                class_stats, 
                "Thống kê tổng quan theo lớp học"
            )
        
        if not trends.empty:
            self.create_styled_worksheet(
                wb, 
                "Xu hướng điểm danh", 
                trends, 
                "Xu hướng điểm danh 30 ngày gần nhất"
            )
        
        if not hourly_pattern.empty:
            self.create_styled_worksheet(
                wb, 
                "Mẫu giờ điểm danh", 
                hourly_pattern, 
                "Phân bố điểm danh theo giờ trong ngày"
            )
        
        # Add charts
        self._add_charts_to_workbook(wb, daily_summary, trends, hourly_pattern)
        
        # Save file
        wb.save(output_path)
        return output_path
    
    def _add_charts_to_workbook(self, workbook, daily_summary: pd.DataFrame, trends: pd.DataFrame, hourly_pattern: pd.DataFrame):
        """Add charts to workbook"""
        try:
            # Create charts sheet
            charts_ws = workbook.create_sheet(title="Biểu đồ")
            
            # Daily attendance chart
            if not daily_summary.empty:
                chart = LineChart()
                chart.title = "Tỷ lệ điểm danh theo ngày"
                chart.y_axis.title = "Tỷ lệ điểm danh (%)"
                chart.x_axis.title = "Ngày"
                
                # Add data to worksheet for chart
                chart_data_start_row = 2
                charts_ws.cell(row=chart_data_start_row, column=1, value="Ngày")
                charts_ws.cell(row=chart_data_start_row, column=2, value="Tỷ lệ điểm danh")
                
                for idx, row in daily_summary.iterrows():
                    charts_ws.cell(row=chart_data_start_row + idx + 1, column=1, value=row['attendance_date'])
                    charts_ws.cell(row=chart_data_start_row + idx + 1, column=2, value=row['attendance_rate'])
                
                data = Reference(charts_ws, min_col=2, min_row=chart_data_start_row, 
                               max_row=chart_data_start_row + len(daily_summary))
                categories = Reference(charts_ws, min_col=1, min_row=chart_data_start_row + 1,
                                     max_row=chart_data_start_row + len(daily_summary))
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                charts_ws.add_chart(chart, "D2")
            
        except Exception as e:
            print(f"Warning: Could not create charts: {e}")
    
    # =============================================================================
    # VISUALIZATION METHODS
    # =============================================================================
    
    def create_attendance_visualization(self, start_date: str, end_date: str) -> Dict[str, str]:
        """
        Create visualization charts and return as base64 encoded images
        
        Returns:
            Dictionary with chart names and base64 encoded image data
        """
        charts = {}
        
        # Set style
        plt.style.use('seaborn-v0_8')
        
        # Daily attendance trend
        trends = self.get_attendance_trends(30)
        if not trends.empty:
            fig, ax = plt.subplots(figsize=(12, 6))
            ax.plot(pd.to_datetime(trends['date']), trends['unique_students'], 
                   marker='o', linewidth=2, markersize=6)
            ax.set_title('Xu hướng số lượng học sinh điểm danh', fontsize=16, fontweight='bold')
            ax.set_xlabel('Ngày', fontsize=12)
            ax.set_ylabel('Số học sinh', fontsize=12)
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            # Convert to base64
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            buffer.seek(0)
            charts['daily_trend'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close()
        
        # Hourly pattern
        hourly = self.get_hourly_attendance_pattern()
        if not hourly.empty:
            fig, ax = plt.subplots(figsize=(10, 6))
            bars = ax.bar(hourly['hour'], hourly['attendance_count'], 
                         color='skyblue', alpha=0.8, edgecolor='navy')
            ax.set_title('Phân bố điểm danh theo giờ', fontsize=16, fontweight='bold')
            ax.set_xlabel('Giờ trong ngày', fontsize=12)
            ax.set_ylabel('Số lần điểm danh', fontsize=12)
            ax.grid(True, alpha=0.3, axis='y')
            
            # Add value labels on bars
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.tight_layout()
            
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            buffer.seek(0)
            charts['hourly_pattern'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close()
        
        # Class attendance comparison
        class_stats = self.get_class_statistics()
        if not class_stats.empty and len(class_stats) > 1:
            fig, ax = plt.subplots(figsize=(10, 6))
            bars = ax.bar(class_stats['class_name'], class_stats['total_students'], 
                         color='lightcoral', alpha=0.8, edgecolor='darkred')
            ax.set_title('So sánh số lượng học sinh theo lớp', fontsize=16, fontweight='bold')
            ax.set_xlabel('Lớp học', fontsize=12)
            ax.set_ylabel('Số học sinh', fontsize=12)
            ax.grid(True, alpha=0.3, axis='y')
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                       f'{int(height)}', ha='center', va='bottom')
            
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            buffer.seek(0)
            charts['class_comparison'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close()
        
        return charts
    
    # =============================================================================
    # UTILITY METHODS
    # =============================================================================
    
    def get_report_summary(self, start_date: str, end_date: str) -> Dict:
        """Get summary statistics for report"""
        summary = {}
        
        # Total records
        attendance_data = self.get_attendance_by_date_range(start_date, end_date)
        summary['total_records'] = len(attendance_data)
        summary['unique_students'] = attendance_data['student_name'].nunique() if not attendance_data.empty else 0
        summary['date_range'] = f"{start_date} đến {end_date}"
        
        # Average confidence
        summary['avg_confidence'] = round(attendance_data['confidence_score'].mean(), 2) if not attendance_data.empty else 0
        
        # Classes involved
        summary['classes_count'] = attendance_data['class_name'].nunique() if not attendance_data.empty else 0
        
        # Most active day
        if not attendance_data.empty:
            daily_counts = attendance_data.groupby(attendance_data['date'].str[:10]).size()
            most_active_day = daily_counts.idxmax()
            summary['most_active_day'] = most_active_day
            summary['most_active_day_count'] = daily_counts.max()
        else:
            summary['most_active_day'] = "N/A"
            summary['most_active_day_count'] = 0
        
        return summary

# Test function for the report generator
def test_report_generator():
    """Test the report generator functionality"""
    import os
    
    # Database path
    db_path = "data/attendance.db"
    
    if not os.path.exists(db_path):
        print(f"❌ Database not found at {db_path}")
        return
    
    # Initialize report generator
    rg = AttendanceReportGenerator(db_path)
    
    # Test date range
    end_date = datetime.now().strftime('%Y-%m-%d')
    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    print(f"🔍 Testing report generator with date range: {start_date} to {end_date}")
    
    try:
        # Test basic queries
        attendance = rg.get_attendance_by_date_range(start_date, end_date)
        print(f"✅ Found {len(attendance)} attendance records")
        
        summary = rg.get_report_summary(start_date, end_date)
        print(f"✅ Report summary: {summary}")
        
        # Test export
        output_file = f"reports/attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs("reports", exist_ok=True)
        
        rg.export_comprehensive_report(start_date, end_date, output_file)
        print(f"✅ Report exported to: {output_file}")
        
    except Exception as e:
        print(f"❌ Error testing report generator: {e}")

if __name__ == "__main__":
    test_report_generator()